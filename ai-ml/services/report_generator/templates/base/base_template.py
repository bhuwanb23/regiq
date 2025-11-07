#!/usr/bin/env python3
"""
REGIQ AI/ML - Base Template System
Abstract base class for all report templates.

This module provides the foundation for all report templates with:
- Template interface definition
- Common functionality
- Data validation
- Error handling
- Output generation

Author: REGIQ AI/ML Team
Version: 1.0.0
"""

import os
import sys
import json
import logging
from abc import ABC, abstractmethod
from pathlib import Path
from typing import Any, Dict, List, Optional, Union, Tuple
from dataclasses import dataclass, asdict
from datetime import datetime
import hashlib

# Add project root to path
sys.path.append(str(Path(__file__).parent.parent.parent.parent.parent))

try:
    import pandas as pd
    import numpy as np
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

from config.env_config import get_env_config

# Configure logging
logger = logging.getLogger(__name__)


@dataclass
class TemplateMetadata:
    """Template metadata information."""
    template_id: str
    template_name: str
    template_type: str
    version: str
    created_at: str
    author: str
    description: str
    supported_formats: List[str]
    
    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary."""
        return asdict(self)


@dataclass 
class ReportSection:
    """Individual report section."""
    section_id: str
    title: str
    content: str
    section_type: str
    order: int
    metadata: Optional[Dict[str, Any]] = None
    
    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary."""
        return asdict(self)


@dataclass
class ReportData:
    """Input data for report generation."""
    regulatory_data: Optional[Dict[str, Any]] = None
    bias_analysis_data: Optional[Dict[str, Any]] = None
    risk_simulation_data: Optional[Dict[str, Any]] = None
    metadata: Optional[Dict[str, Any]] = None
    
    def validate(self) -> Tuple[bool, List[str]]:
        """Validate input data."""
        errors = []
        
        # Check if at least one data source is provided
        if not any([self.regulatory_data, self.bias_analysis_data, self.risk_simulation_data]):
            errors.append("At least one data source must be provided")
        
        # Validate individual data sources
        if self.regulatory_data and not isinstance(self.regulatory_data, dict):
            errors.append("Regulatory data must be a dictionary")
            
        if self.bias_analysis_data and not isinstance(self.bias_analysis_data, dict):
            errors.append("Bias analysis data must be a dictionary")
            
        if self.risk_simulation_data and not isinstance(self.risk_simulation_data, dict):
            errors.append("Risk simulation data must be a dictionary")
        
        return len(errors) == 0, errors


class BaseTemplate(ABC):
    """
    Abstract base class for all report templates.
    
    Provides common functionality and interface for:
    - Executive Reports
    - Technical Reports  
    - Regulatory Reports
    """
    
    def __init__(self, template_id: str, template_name: str, version: str = "1.0.0"):
        """
        Initialize base template.
        
        Args:
            template_id: Unique template identifier
            template_name: Human-readable template name
            version: Template version
        """
        self.template_id = template_id
        self.template_name = template_name
        self.version = version
        self.created_at = datetime.utcnow().isoformat()
        self.sections: List[ReportSection] = []
        self.metadata = TemplateMetadata(
            template_id=template_id,
            template_name=template_name,
            template_type=self.__class__.__name__,
            version=version,
            created_at=self.created_at,
            author="REGIQ AI/ML Team",
            description=self.get_description(),
            supported_formats=self.get_supported_formats()
        )
        
        # Initialize logger
        self.logger = logging.getLogger(f"{__name__}.{self.__class__.__name__}")
        
    @abstractmethod
    def get_description(self) -> str:
        """Get template description."""
        pass
    
    @abstractmethod
    def get_supported_formats(self) -> List[str]:
        """Get supported output formats."""
        pass
    
    @abstractmethod
    def generate_sections(self, data: ReportData) -> List[ReportSection]:
        """Generate report sections from input data."""
        pass
    
    @abstractmethod
    def validate_data(self, data: ReportData) -> Tuple[bool, List[str]]:
        """Validate input data for this template."""
        pass
    
    def add_section(self, section: ReportSection) -> None:
        """Add a section to the report."""
        self.sections.append(section)
        self.logger.debug(f"Added section: {section.section_id}")
    
    def remove_section(self, section_id: str) -> bool:
        """Remove a section by ID."""
        initial_count = len(self.sections)
        self.sections = [s for s in self.sections if s.section_id != section_id]
        removed = len(self.sections) < initial_count
        
        if removed:
            self.logger.debug(f"Removed section: {section_id}")
        else:
            self.logger.warning(f"Section not found: {section_id}")
            
        return removed
    
    def get_section(self, section_id: str) -> Optional[ReportSection]:
        """Get a section by ID."""
        for section in self.sections:
            if section.section_id == section_id:
                return section
        return None
    
    def reorder_sections(self) -> None:
        """Reorder sections by their order attribute."""
        self.sections.sort(key=lambda x: x.order)
        self.logger.debug("Sections reordered")
    
    def generate_report(self, data: ReportData, output_format: str = "html") -> Dict[str, Any]:
        """
        Generate complete report from input data.
        
        Args:
            data: Input data for report generation
            output_format: Output format (html, pdf, json)
            
        Returns:
            Generated report data
        """
        try:
            # Validate input data
            is_valid, errors = self.validate_data(data)
            if not is_valid:
                raise ValueError(f"Data validation failed: {', '.join(errors)}")
            
            # Generate sections
            self.logger.info(f"Generating {self.template_name} report")
            sections = self.generate_sections(data)
            
            # Clear existing sections and add new ones
            self.sections.clear()
            for section in sections:
                self.add_section(section)
            
            # Reorder sections
            self.reorder_sections()
            
            # Generate output based on format
            if output_format.lower() == "html":
                output = self._generate_html_output()
            elif output_format.lower() == "pdf":
                output = self._generate_pdf_output()
            elif output_format.lower() == "json":
                output = self._generate_json_output()
            elif output_format.lower() == "csv":
                output = self._generate_csv_output()
            elif output_format.lower() in ["xlsx", "excel"]:
                output = self._generate_excel_output()
            else:
                raise ValueError(f"Unsupported output format: {output_format}")
            
            # Add metadata
            result = {
                "template_metadata": self.metadata.to_dict(),
                "generation_timestamp": datetime.utcnow().isoformat(),
                "output_format": output_format,
                "sections_count": len(self.sections),
                "content": output
            }
            
            self.logger.info(f"Report generated successfully: {len(self.sections)} sections")
            return result
            
        except Exception as e:
            self.logger.error(f"Report generation failed: {str(e)}")
            raise
    
    def _generate_html_output(self) -> str:
        """Generate HTML output."""
        html_parts = []
        
        # HTML header
        html_parts.append(f"""
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>{self.template_name} - REGIQ Report</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 40px; line-height: 1.6; }}
                .header {{ border-bottom: 2px solid #333; padding-bottom: 20px; margin-bottom: 30px; }}
                .section {{ margin-bottom: 30px; }}
                .section-title {{ color: #333; border-left: 4px solid #007acc; padding-left: 15px; }}
                .metadata {{ background: #f5f5f5; padding: 15px; border-radius: 5px; margin-bottom: 20px; }}
                .timestamp {{ color: #666; font-size: 0.9em; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>{self.template_name}</h1>
                <div class="metadata">
                    <p><strong>Template ID:</strong> {self.template_id}</p>
                    <p><strong>Version:</strong> {self.version}</p>
                    <p class="timestamp"><strong>Generated:</strong> {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}</p>
                </div>
            </div>
        """)
        
        # Sections
        for section in self.sections:
            html_parts.append(f"""
            <div class="section">
                <h2 class="section-title">{section.title}</h2>
                <div class="section-content">
                    {section.content}
                </div>
            </div>
            """)
        
        # HTML footer
        html_parts.append("""
            <div class="footer" style="margin-top: 50px; padding-top: 20px; border-top: 1px solid #ccc; text-align: center; color: #666;">
                <p>Generated by REGIQ AI/ML Report System</p>
            </div>
        </body>
        </html>
        """)
        
        return "".join(html_parts)
    
    def _generate_pdf_output(self) -> bytes:
        """Generate PDF output using ReportLab."""
        try:
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib.units import inch
            from reportlab.lib import colors
            from reportlab.pdfgen import canvas
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
            from io import BytesIO
            import re
            
            # Create PDF in memory
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
            styles = getSampleStyleSheet()
            
            # Create custom styles
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                spaceAfter=30,
                alignment=TA_CENTER,
                textColor=colors.darkblue
            )
            
            section_title_style = ParagraphStyle(
                'SectionTitle',
                parent=styles['Heading2'],
                fontSize=16,
                spaceAfter=12,
                textColor=colors.blue
            )
            
            normal_style = styles['Normal']
            normal_style.fontSize = 10
            normal_style.spaceAfter = 6
            
            # Build content
            story = []
            
            # Add title
            title = Paragraph(self.template_name, title_style)
            story.append(title)
            story.append(Spacer(1, 0.3*inch))
            
            # Add metadata
            metadata_text = f"Template ID: {self.template_id} | Version: {self.version} | Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}"
            metadata = Paragraph(metadata_text, styles['Italic'])
            story.append(metadata)
            story.append(Spacer(1, 0.3*inch))
            
            # Add sections
            for section in self.sections:
                # Section title
                section_title = Paragraph(section.title, section_title_style)
                story.append(section_title)
                story.append(Spacer(1, 0.1*inch))
                
                # Section content - clean HTML tags
                clean_content = re.sub('<[^<]+?>', '', section.content).strip()
                if clean_content:
                    content = Paragraph(clean_content, normal_style)
                    story.append(content)
                    story.append(Spacer(1, 0.2*inch))
            
            # Add footer
            story.append(Spacer(1, 0.5*inch))
            footer_text = "Generated by REGIQ AI/ML Report System"
            footer = Paragraph(footer_text, styles['Italic'])
            story.append(footer)
            
            # Build PDF with custom header/footer
            def add_header_footer(canvas, doc):
                canvas.saveState()
                # Header
                canvas.setFont('Helvetica-Bold', 10)
                canvas.drawString(inch, 10.5*inch, f"{self.template_name} - {self.template_id}")
                # Footer
                canvas.setFont('Helvetica', 8)
                canvas.drawRightString(7.5*inch, 0.5*inch, f"Page {canvas.getPageNumber()}")
                canvas.drawString(inch, 0.5*inch, datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC'))
                canvas.restoreState()
            
            # Build PDF
            doc.build(story, onFirstPage=add_header_footer, onLaterPages=add_header_footer)
            
            # Return PDF bytes
            buffer.seek(0)
            return buffer.getvalue()
            
        except ImportError:
            self.logger.error("ReportLab not installed. Cannot generate PDF output.")
            raise ImportError("ReportLab library is required for PDF export. Please install it with: pip install reportlab")
        except Exception as e:
            self.logger.error(f"Failed to generate PDF output: {str(e)}")
            # Fallback to HTML if PDF generation fails
            html_content = self._generate_html_output()
            # Convert string to bytes
            return html_content.encode('utf-8')
    
    def _generate_json_output(self) -> Dict[str, Any]:
        """Generate JSON output."""
        return {
            "template_metadata": self.metadata.to_dict(),
            "sections": [section.to_dict() for section in self.sections],
            "generation_info": {
                "timestamp": datetime.utcnow().isoformat(),
                "sections_count": len(self.sections)
            }
        }
    
    def _generate_csv_output(self) -> str:
        """Generate CSV output."""
        import csv
        from io import StringIO
        
        output = StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow(['Section ID', 'Title', 'Content', 'Type', 'Order'])
        
        # Write sections
        for section in self.sections:
            # Clean content for CSV (remove HTML tags if needed)
            clean_content = section.content
            if clean_content.startswith('<') and clean_content.endswith('>'):
                # Simple HTML tag removal
                import re
                clean_content = re.sub('<[^<]+?>', '', clean_content)
            
            writer.writerow([
                section.section_id,
                section.title,
                clean_content,
                section.section_type,
                section.order
            ])
        
        return output.getvalue()
    
    def _generate_excel_output(self) -> bytes:
        """Generate Excel output using openpyxl."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            from io import BytesIO
            
            wb = Workbook()
            ws = wb.active
            if ws is not None:
                ws.title = "Report Data"
                
                # Add headers
                headers = ['Section ID', 'Title', 'Content', 'Type', 'Order']
                ws.append(headers)
                
                # Style headers
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Add data
                for section in self.sections:
                    # Clean content for Excel (remove HTML tags if needed)
                    clean_content = section.content
                    if clean_content.startswith('<') and clean_content.endswith('>'):
                        # Simple HTML tag removal
                        import re
                        clean_content = re.sub('<[^<]+?>', '', clean_content)
                    
                    ws.append([
                        section.section_id,
                        section.title,
                        clean_content,
                        section.section_type,
                        section.order
                    ])
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    if column and len(column) > 0 and column[0].column is not None:
                        column_letter = get_column_letter(column[0].column)
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length + 2)
                        if ws.column_dimensions and column_letter:
                            ws.column_dimensions[column_letter].width = min(adjusted_width, 50)
                
                # Save to bytes
                buffer = BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                return buffer.getvalue()
            else:
                raise Exception("Failed to create Excel worksheet")
            
        except ImportError:
            self.logger.error("openpyxl not installed. Cannot generate Excel output.")
            raise ImportError("openpyxl library is required for Excel export. Please install it with: pip install openpyxl")
        except Exception as e:
            self.logger.error(f"Failed to generate Excel output: {str(e)}")
            raise
    
    def save_report(self, data: ReportData, output_path: str, output_format: str = "html") -> str:
        """
        Generate and save report to file.
        
        Args:
            data: Input data for report generation
            output_path: Output file path
            output_format: Output format
            
        Returns:
            Path to saved file
        """
        try:
            # Generate report
            report = self.generate_report(data, output_format)
            
            # Ensure output directory exists
            output_file = Path(output_path)
            output_file.parent.mkdir(parents=True, exist_ok=True)
            
            # Save based on format
            if output_format.lower() == "json":
                with open(output_file, 'w', encoding='utf-8') as f:
                    json.dump(report, f, indent=2, ensure_ascii=False)
            elif output_format.lower() == "csv":
                # CSV content is text
                with open(output_file, 'w', encoding='utf-8', newline='') as f:
                    f.write(report["content"])
            elif output_format.lower() in ["xlsx", "excel", "pdf"]:
                # Excel and PDF content are binary
                with open(output_file, 'wb') as f:
                    f.write(report["content"])
            else:
                # HTML or other text formats
                with open(output_file, 'w', encoding='utf-8') as f:
                    f.write(report["content"])
            
            self.logger.info(f"Report saved to: {output_file}")
            return str(output_file)
            
            self.logger.info(f"Report saved to: {output_file}")
            return str(output_file)
            
        except Exception as e:
            self.logger.error(f"Failed to save report: {str(e)}")
            raise
    
    def get_template_info(self) -> Dict[str, Any]:
        """Get template information."""
        return {
            "metadata": self.metadata.to_dict(),
            "sections_count": len(self.sections),
            "current_sections": [
                {"id": s.section_id, "title": s.title, "order": s.order} 
                for s in self.sections
            ]
        }
    
    def __str__(self) -> str:
        """String representation."""
        return f"{self.__class__.__name__}(id={self.template_id}, name={self.template_name})"
    
    def __repr__(self) -> str:
        """Detailed representation."""
        return (f"{self.__class__.__name__}(template_id='{self.template_id}', "
                f"template_name='{self.template_name}', version='{self.version}', "
                f"sections={len(self.sections)})")
